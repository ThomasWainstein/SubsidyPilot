"""Pure Python document extraction utilities.

This module provides a light‑weight implementation of the
``python_document_extractor`` used across several AgriTool
pipelines.  It supports PDF, DOCX and XLSX documents without any
external services.
"""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any, Dict


class DocumentExtractionResult:
    """Container for document extraction results."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        self.extraction_method: str | None = None
        self.text_content: str = ""
        self.metadata: Dict[str, Any] = {}
        self.success: bool = False
        self.error: str | None = None
        self.character_count: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "file_path": self.file_path,
            "file_name": self.file_name,
            "file_size": self.file_size,
            "extraction_method": self.extraction_method,
            "text_content": self.text_content,
            "metadata": self.metadata,
            "success": self.success,
            "error": self.error,
            "character_count": self.character_count,
        }


class PythonDocumentExtractor:
    """Simple pure Python document extractor."""

    def __init__(
        self,
        enable_ocr: bool = True,
        ocr_language: str = "eng",
        max_file_size_mb: float = 25.0,
    ) -> None:
        # The parameters are accepted for backwards compatibility.
        self.enable_ocr = enable_ocr
        self.ocr_language = ocr_language
        self.max_file_size_mb = max_file_size_mb
        self.max_file_size_bytes = int(max_file_size_mb * 1024 * 1024)
        self.logger = logging.getLogger(__name__)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def extract_document(self, file_path: str) -> DocumentExtractionResult:
        """Extract text from the given document."""

        result = DocumentExtractionResult(file_path)
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(file_path)

            if os.path.getsize(file_path) > self.max_file_size_bytes:
                raise ValueError("File too large")

            ext = Path(file_path).suffix.lower()
            if ext == ".pdf":
                text = self._extract_pdf(file_path)
                result.extraction_method = "pdf"
            elif ext in {".doc", ".docx"}:
                text = self._extract_docx(file_path)
                result.extraction_method = "docx"
            elif ext in {".xls", ".xlsx"}:
                text = self._extract_xlsx(file_path)
                result.extraction_method = "xlsx"
            else:
                raise ValueError(f"Unsupported file type: {ext}")

            result.text_content = text
            result.success = True
            result.character_count = len(text)
        except Exception as exc:  # pragma: no cover - defensive
            result.error = str(exc)
            self.logger.error("Document extraction failed", exc_info=exc)
        return result

    # ------------------------------------------------------------------
    # Format specific helpers
    # ------------------------------------------------------------------
    def _extract_pdf(self, file_path: str) -> str:
        import pdfplumber

        with pdfplumber.open(file_path) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
        return "\n".join(pages)

    def _extract_docx(self, file_path: str) -> str:
        from docx import Document

        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs)

    def _extract_xlsx(self, file_path: str) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        texts: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(c) for c in row if c is not None)
                if row_text:
                    texts.append(row_text)
        return "\n".join(texts)


def extract_document_text(file_path: str) -> str:
    """Convenience helper returning only the extracted text."""

    extractor = PythonDocumentExtractor()
    return extractor.extract_document(file_path).text_content


class ScraperDocumentExtractor:
    """Wrapper used by the scraper pipeline."""

    def __init__(self, enable_ocr: bool = True, max_file_size_mb: float = 25.0) -> None:
        self.extractor = PythonDocumentExtractor(
            enable_ocr=enable_ocr,
            max_file_size_mb=max_file_size_mb,
        )
        self.logger = logging.getLogger(__name__)

    def extract_attachment_text(self, attachment_path: str) -> Dict[str, Any]:
        try:
            result = self.extractor.extract_document(attachment_path)
            data = result.to_dict()
            data["extracted_for_scraper"] = True
            data["attachment_path"] = attachment_path
            return data
        except Exception as exc:  # pragma: no cover - defensive
            self.logger.error("Document extraction failed", exc_info=exc)
            return {
                "file_path": attachment_path,
                "file_name": os.path.basename(attachment_path),
                "success": False,
                "error": str(exc),
                "text_content": "",
                "extraction_method": "failed",
                "extracted_for_scraper": True,
            }
