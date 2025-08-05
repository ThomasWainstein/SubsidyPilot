"""
XLSX document extractor for spreadsheet-based forms.
"""

import asyncio
import logging
from typing import Dict, Any

from openpyxl import load_workbook


class XLSXExtractor:
    """XLSX document schema extractor."""
    
    def __init__(self, logger: logging.Logger):
        self.logger = logger
    
    async def extract_schema(self, file_path: str) -> Dict[str, Any]:
        """
        Extract schema from XLSX document.
        
        Args:
            file_path: Path to XLSX file.
            
        Returns:
            Extracted schema data with metadata.
        """
        self.logger.info("Starting XLSX extraction", extra={
            "file_path": file_path,
            "extractor": "XLSXExtractor"
        })
        
        # Run extraction in thread pool
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._extract_xlsx_sync, file_path)
    
    def _extract_xlsx_sync(self, file_path: str) -> Dict[str, Any]:
        """Synchronous XLSX extraction."""
        fields = []
        raw_unclassified = []
        metadata = {
            "method": "xlsx_extraction",
            "extractor": "XLSXExtractor",
            "worksheet_count": 0,
            "worksheets_processed": []
        }
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            metadata["worksheet_count"] = len(workbook.sheetnames)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                metadata["worksheets_processed"].append(sheet_name)
                
                self.logger.debug(f"Processing worksheet: {sheet_name}")
                
                # Extract fields from worksheet
                sheet_data = self._extract_sheet_fields(sheet, sheet_name)
                fields.extend(sheet_data['fields'])
                raw_unclassified.extend(sheet_data['raw_unclassified'])
            
            workbook.close()
            
        except Exception as e:
            self.logger.error("XLSX extraction failed", extra={
                "file_path": file_path,
                "error": str(e),
                "error_type": type(e).__name__
            })
            raw_unclassified.append(f"XLSX extraction error: {str(e)}")
        
        return {
            "fields": fields,
            "raw_unclassified": raw_unclassified,
            "metadata": metadata
        }
    
    def _extract_sheet_fields(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extract fields from a single worksheet."""
        fields = []
        raw_unclassified = []
        
        from utils.field_parser import FieldParser
        parser = FieldParser(self.logger)
        
        # Track potential field headers and data validation
        headers = {}
        validation_rules = {}
        
        # Scan for form-like patterns in cells
        for row_idx, row in enumerate(sheet.iter_rows(max_col=20, max_row=200), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    
                    # Check for field-like patterns
                    field_data = self._parse_cell_for_field(cell_value, sheet_name, row_idx, col_idx)
                    if field_data:
                        # Check for data validation on adjacent cells
                        validation = self._check_data_validation(sheet, row_idx, col_idx + 1)
                        if validation:
                            field_data.update(validation)
                        
                        fields.append(field_data)
                    elif len(cell_value) > 200:  # Long text that might be instructions
                        raw_unclassified.append(cell_value)
                    
                    # Track potential headers for context
                    if ':' not in cell_value and len(cell_value) < 50:
                        headers[f"{row_idx}_{col_idx}"] = cell_value
        
        # Look for structured forms based on headers
        structured_fields = self._detect_structured_forms(sheet, headers)
        fields.extend(structured_fields)
        
        return {
            "fields": fields,
            "raw_unclassified": raw_unclassified
        }
    
    def _parse_cell_for_field(self, cell_value: str, sheet_name: str, row: int, col: int) -> Dict[str, Any]:
        """Parse Excel cell value to identify if it's a form field."""
        from utils.field_parser import FieldParser
        parser = FieldParser(self.logger)
        
        if ':' in cell_value:
            parts = cell_value.split(':', 1)
            label = parts[0].strip()
            help_text = parts[1].strip() if len(parts) > 1 else ""
            
            if len(label) > 2 and len(label) < 150:
                return {
                    "name": parser.normalize_field_name(label),
                    "label": label,
                    "type": parser.infer_field_type(label),
                    "required": parser.is_field_required(label),
                    "help": help_text,
                    "source": f"sheet_{sheet_name}_cell_{row}_{col}"
                }
        
        # Check for question patterns
        if '?' in cell_value and len(cell_value) < 200:
            return {
                "name": parser.normalize_field_name(cell_value),
                "label": cell_value,
                "type": parser.infer_field_type(cell_value),
                "required": parser.is_field_required(cell_value),
                "help": "",
                "source": f"sheet_{sheet_name}_cell_{row}_{col}"
            }
        
        return None
    
    def _check_data_validation(self, sheet, row: int, col: int) -> Dict[str, Any]:
        """Check for data validation rules on a cell."""
        try:
            cell = sheet.cell(row=row, column=col)
            if hasattr(cell, 'data_validation') and cell.data_validation:
                validation = cell.data_validation
                validation_info = {}
                
                if hasattr(validation, 'type') and validation.type:
                    validation_info['validation_type'] = validation.type
                
                if hasattr(validation, 'formula1') and validation.formula1:
                    # Extract list options if it's a list validation
                    if validation.type == 'list':
                        options = str(validation.formula1).strip('"').split(',')
                        validation_info['options'] = [opt.strip() for opt in options]
                        validation_info['type'] = 'select'
                
                return validation_info
                
        except Exception:
            pass
        
        return {}
    
    def _detect_structured_forms(self, sheet, headers: Dict[str, str]) -> List[Dict[str, Any]]:
        """Detect structured form patterns in the spreadsheet."""
        fields = []
        
        # Look for common form patterns like header rows followed by input rows
        # This could be enhanced based on actual form structures found
        
        return fields